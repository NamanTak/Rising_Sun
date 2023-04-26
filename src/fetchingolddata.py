import time
from datetime import datetime
import access_token
import repository
import pandas as pd
import numpy as np
import openpyxl

# STOCK_NAME = ["TECHM","BHARTIARTL","POWERGRID","HCLTECH",
# "NTPC","INFY","HINDALCO","ULTRACEMCO","EICHERMOT","COALINDIA","SBILIFE","HDFCLIFE","BAJAJ-AUTO","BAJAJFINSV","LT","HINDUNILVR","JSWSTEEL","ITC","TATASTEEL","M&M","SUNPHARMA","INDUSINDBK","APOLLOHOSP","DRREDDY","BAJFINANCE","HDFCBANK","TATAMOTORS","TCS","DIVISLAB","HEROMOTOCO","ASIANPAINT","ONGC","GRASIM","AXISBANK","HDFC","UPL","KOTAKBANK","NESTLEIND","TITAN","ICICIBANK","TATACONSUM","MARUTI","BPCL","BRITANNIA","ADANIPORTS","CIPLA","ADANIENT"]
STOCK_NAME = ["MARUTI"]
RESOLUTION = "1"

# STOCK_NAME = "HINDUNILVR"
# RESOLUTION = "5"



worksheet = ["Sheet"]

FROM_DAYS = 100
FUNDS = 250000 * 5
rr_ratio = 4

dema_time_period = 3
supertrend_ATR = 12
supertrend_Multiplier = 3
per = 1.03  # sell percentage criteria

# Constants
BUY_SIDE = 1
SELL_SIDE = -1

# Variables
OPEN_VALUES = []
HIGH_VALUES = []
LOW_VALUES = []
CLOSE_VALUES = []
EPOCH_VALUES = []
EMA_VALUES = []
FLAG = 0

# orderId = ""
# quantity = 1
# type = 1
# product_type = "INTRADAY"
# validity = "DAY"
# symbol = "NSE:" + STOCK_NAME + "-EQ"
# limit_price = 0

# ***********************************************************************************************************************

def pass_stock_data(stock_name: str, resolution: str,a,b,date_format="0" ):
    """
        Returns the dictionary(key-value pair)
        :param stock_name:
        :param resolution: Time frame of a chart i.e., 5min, 10min, 15min, etc.
        :param from_days: Data lene ke liye defined days pehle se
        :param to_days: Current Time is already defined
        :param cont_flag: Flag need for FYERS API
        :param date_format: As defined in the FYERS API
        :return:data
        """
    print("\nGenerating String of Stock Meta Data")
    data = {"symbol": f"NSE:{stock_name}-EQ", "resolution": f"{resolution}", "date_format": f"{date_format}",
            "range_from": a,
            "range_to": b, "cont_flag": "1"}
    print("Generated String of Stock Meta Data is - \n", data, "\n")
    return data

a = 1483213261
b = 8640000 + a
currenttime = 1676959335
count = 0

print("Writing Started")
now = datetime.now()

current_time_before_loop = now.strftime("%H:%M:%S")
print("Start Time =", current_time_before_loop)

stock_name_length = len(STOCK_NAME)

for i in range (stock_name_length):

    openpyxl.Workbook().save(f"{STOCK_NAME[i]}.xlsx")

    # Load the existing workbook
    workbook = openpyxl.load_workbook(f'{STOCK_NAME[i]}.xlsx')

    # Save the workbook
    workbook.save(f'{STOCK_NAME[i]}.xlsx')


    while(a < currenttime):
        print("Start Time =", current_time_before_loop)

        # Generating Stock Details
        stock_meta_data = pass_stock_data(STOCK_NAME[i], RESOLUTION,a,b)

        # Getting the entire data of the stock
        entire_stock_data = access_token.get_fyers_entry_point().history(stock_meta_data)

        # length of entire stock data
        length_of_entire_stock_data = len(entire_stock_data['candles'])

        a = b
        b = b + 8640000

        # read the demo2.xlsx file
        df = pd.read_excel(f"{STOCK_NAME[i]}.xlsx")

        # appending the data of df after the data of demo1.xlsx
        with pd.ExcelWriter(f"{STOCK_NAME[i]}.xlsx",mode="a",engine="openpyxl",if_sheet_exists="overlay") as writer:
            pd.DataFrame(entire_stock_data).to_excel(writer, sheet_name=f"{worksheet[i]}",header=None, startrow=writer.sheets[f"{worksheet[i]}"].max_row,index=False)

        # with pd.ExcelWriter(f'{STOCK_NAME}.xlsx', mode='a',engine="openpyxl",if_sheet_exists="overlay") as writer:
        #     pd.DataFrame(entire_stock_data).to_excel(writer, sheet_name='sheet',header=None, startrow=writer.sheets['sheet'
        #                                                                                                             ''].max_row,index=False)

        print(f"For the {count + 1} time ")

        count = count + 1

        #if we are running the code for multiple stocks at a time then this values should be reset to year 2017
        # a = 1483213261
        # b = 8640000 + a

now = datetime.now()
current_time_after_loop = now.strftime("%H:%M:%S")
print("Current Time =", current_time_after_loop)

print(f"Total time taken by loop = {repository.calcTime(current_time_before_loop,current_time_after_loop)}")
